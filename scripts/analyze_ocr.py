import json
import re
from openpyxl import Workbook

# Load the letters data
with open('C:/dev/Site SC/data/cartas.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Base URL for the website
base_url = "https://aleapc.github.io/minhas-cartas/cartas.html"

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "OCR Problems"

# Headers
ws['A1'] = "Link da Carta"
ws['B1'] = "ID"
ws['C1'] = "Pagina"
ws['D1'] = "Texto com Problema (OCR)"
ws['E1'] = "Texto Corrigido (preencher)"

# Style headers
for col in ['A', 'B', 'C', 'D', 'E']:
    ws[f'{col}1'].font = ws[f'{col}1'].font.copy(bold=True)

# Function to detect OCR problems
def has_ocr_problems(text):
    if not text or len(text) < 20:
        return True, "Texto muito curto ou vazio"

    # Check for excessive special characters
    special_chars = len(re.findall(r'[|=\[\]{}]', text))
    if special_chars > len(text) * 0.1:
        return True, "Muitos caracteres especiais"

    # Check for fragmented text (many single characters separated)
    fragments = len(re.findall(r'\b[A-Za-z]\b', text))
    if fragments > len(text) * 0.05:
        return True, "Texto fragmentado"

    # Check for gibberish patterns
    gibberish = len(re.findall(r'[A-Z]{2,}\s+[A-Z]{2,}', text))
    if gibberish > 5:
        return True, "Padroes de OCR incorreto"

    # Check for excessive line breaks with little content
    lines = text.split('\n')
    short_lines = sum(1 for line in lines if len(line.strip()) < 5 and line.strip())
    if short_lines > len(lines) * 0.5:
        return True, "Muitas linhas curtas"

    return False, ""

# Analyze each letter
row = 2
problem_count = 0

for carta in data['cartas']:
    texto = carta.get('texto', '')
    has_problem, problem_type = has_ocr_problems(texto)

    if has_problem:
        problem_count += 1
        link = f"{base_url}#{carta['id']}"

        ws[f'A{row}'] = link
        ws[f'B{row}'] = carta['id']
        ws[f'C{row}'] = carta.get('pagina', '')
        ws[f'D{row}'] = texto[:2000] if texto else "(vazio)"  # Limit text length
        ws[f'E{row}'] = ""

        row += 1

# Adjust column widths
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 80
ws.column_dimensions['E'].width = 80

# Save the workbook
output_path = 'C:/dev/Site SC/ocr_problems.xlsx'
wb.save(output_path)

print(f"Total de cartas analisadas: {len(data['cartas'])}")
print(f"Cartas com problemas de OCR: {problem_count}")
print(f"Arquivo salvo em: {output_path}")
